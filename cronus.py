"""
Aplicação de Exportação de Dados do BCB para Excel

Esta aplicação realiza o processamento de dados provenientes de uma API pública do Banco Central do Brasil,
relacionada à agenda das autoridades, está aplicação filtra informações específicas e exporta esses dados para um arquivo Excel.

Funcionamento:
- Recebe os dados da agenda de uma API pública do Banco Central do Brasil.
- Processa os dados e os filtras conforme a necessidade do usuário de querer vê apenas determinado cargo.
- Cria um DataFrame utilizando a biblioteca Pandas para estruturar os dados.
- Exporta os dados para um arquivo Excel (.xlsx), utilizando o nome "YYYY-MM-DD_YYYY-MM-DD.xlsx".
- Formata o arquivo Excel com bordas médias em todas as células e um fundo amarelo,verde e vermelho claro agradável aos olhos humanos.
- Exibe informações adicionais no console, como o nome do arquivo exportado, número de linhas e tamanho do arquivo.

Este código pode ser adaptado para diferentes conjuntos de dados e necessidades de formatação de Excel,
facilitando a exportação de informações estruturadas para relatórios ou análises.

Desenvolvido por Matheus, também conhecido como blkz.
Github: https://github.com/blkzy
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from os import path

import pandas as pd
import requests
import filters
import process
import utils

BANNER = """   __________  ____  _   ____  _______
  / ____/ __ \/ __ \/ | / / / / / ___/
 / /   / /_/ / / / /  |/ / / / /\__ \ 
/ /___/ _, _/ /_/ / /|  / /_/ /___/ /  v0.1
\____/_/ |_|\____/_/ |_/\____//____/   github: blkzy
    Agenda de Autoridades do BCB
"""

MANUAL = """-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
COR VERDE: Há evento(s) aberto(s) à imprensa;
COR AMARELA: Há evento(s) aberto(s) e também fechados à imprensa;
COR VERMELHA: Há evento(s) fechados à imprensa.
-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
"""

def request_schedule_data(from_date, to_date):
    response = requests.get(f"https://www.bcb.gov.br/api/servico/sitebcb/agendadiretoria?lista=Agenda%20da%20Diretoria&inicioAgenda=%27{from_date}%27&fimAgenda=%27{to_date}%27")

    return response.json()

def init():
    print(BANNER)
    print(MANUAL)
    
def main():
    response_api_json = request_schedule_data("2023-02-28", "2024-06-28")

    process_informations = process.process_informations(response_api_json)
    informations_filter = filters.filter_by_role_name(process_informations, "Presidente")
        
    data_frame = [
        [
            info["date_of_event"],
            info["meeting_subject"],
            info["meeting_location"],
            info["role"],
            info["organization"]
        ]
        for info in informations_filter
    ]

    df = pd.DataFrame(data_frame, columns=["Data do Evento", "Assunto do Evento", "Local do Evento", "Cargo", "Orgão"])

    file_name = "Agenda - 2023-02-28_2024-06-28.xlsx"
    df.to_excel(file_name, index=False)

    workbook = load_workbook(file_name)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

        if any("(aberto à imprensa" in str(cell.value) for cell in row):
            for cell in row:
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                
        if any("(fechado à imprensa" in str(cell.value) for cell in row):
            for cell in row:
                cell.fill = fill_close = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        if any("(aberto à imprensa" in str(cell.value) for cell in row) and any("(fechado à imprensa" in str(cell.value) for cell in row):
            for cell in row:
                cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

    workbook.save(file_name)

    print("[+] Arquivo exportado para Excel com sucesso.\n")
    print(f"[i] Nome do arquivo: '{file_name}'")
    print(f"[i] Número de linhas: {len(df)+1}")
    print(f"[i] Tamanho do arquivo: {utils.file_size(path.getsize(file_name))}")
    
if __name__ == "__main__":
    init()
    main()